from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database.connection import get_db
from app.models.swimming_pool import SwimmingPool
import io
from typing import List
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter(prefix="/excel", tags=["Excel Operations"])

@router.get("/export")
def export_pools_to_excel(db: Session = Depends(get_db)):
    """
    모든 수영장 정보를 Excel 파일(.xlsx)로 다운로드

    반환되는 Excel 형식:
    - ID, 수영장명, 주소, 전화번호, 일일권, 자유수영, 웹사이트, 비고
    """
    # 모든 수영장 조회
    pools = db.query(SwimmingPool).order_by(SwimmingPool.id).all()

    # Excel 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "수영장 정보"

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 헤더 작성
    headers = ['ID', '수영장명', '주소', '전화번호', '한달 수강권', '자유수영', '웹사이트', '비고']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 데이터 작성
    for row_num, pool in enumerate(pools, 2):
        ws.cell(row=row_num, column=1, value=pool.id)
        ws.cell(row=row_num, column=2, value=pool.name)
        ws.cell(row=row_num, column=3, value=pool.address)
        ws.cell(row=row_num, column=4, value=pool.phone or '')
        ws.cell(row=row_num, column=5, value=pool.monthly_lesson_price or '')
        ws.cell(row=row_num, column=6, value=pool.free_swim_price or '')
        ws.cell(row=row_num, column=7, value=pool.url or '')
        ws.cell(row=row_num, column=8, value='')  # 비고

    # 컬럼 너비 자동 조정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 20

    # 메모리에 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Excel 파일로 반환
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=swimming_pools.xlsx"
        }
    )


@router.post("/import")
async def import_pools_from_excel(
    file: UploadFile = File(..., description="Excel 파일 (.xlsx)"),
    db: Session = Depends(get_db)
):
    """
    Excel 파일에서 수영장 정보를 읽어서 DB 업데이트

    Excel 형식:
    - ID, 수영장명, 주소, 전화번호, 한달 수강권, 자유수영, 웹사이트, 비고

    업데이트 가능한 필드:
    - 주소 (address)
    - 전화번호 (phone)
    - 한달 수강권 (monthly_lesson_price) - 숫자 또는 문자열 (예: "150000" 또는 "가격 다양, 표 참조")
    - 자유수영 (free_swim_price) - 숫자 또는 문자열 (예: "8000" 또는 "시간대별 상이")
    - 웹사이트 (url)

    주의:
    - ID가 있는 행만 업데이트 (새로운 수영장 추가는 불가)
    - 비어있는 칸은 업데이트하지 않음
    """
    # 파일 확장자 체크
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Excel 파일(.xlsx)만 업로드 가능합니다")

    try:
        # 파일 읽기
        contents = await file.read()

        # Excel 파일 로드
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active

        updated_count = 0
        error_rows = []

        # 헤더 행 읽기 (첫 번째 행)
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        # 데이터 행 처리 (2번째 행부터)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # 행을 딕셔너리로 변환
                row_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        row_dict[header] = row[idx]

                # ID 체크
                if 'ID' not in row_dict or not row_dict['ID']:
                    continue

                pool_id = int(row_dict['ID'])

                # DB에서 수영장 찾기
                pool = db.query(SwimmingPool).filter(SwimmingPool.id == pool_id).first()
                if not pool:
                    error_rows.append({
                        "row": row_num,
                        "id": pool_id,
                        "error": "수영장을 찾을 수 없음"
                    })
                    continue

                # 업데이트할 필드 체크
                updated = False

                # 한달 수강권 가격 (문자열로 저장)
                if '한달 수강권' in row_dict and row_dict['한달 수강권']:
                    value = str(row_dict['한달 수강권']).strip()
                    if value:
                        pool.monthly_lesson_price = value
                        updated = True

                # 자유수영 가격 (문자열로 저장)
                if '자유수영' in row_dict and row_dict['자유수영']:
                    value = str(row_dict['자유수영']).strip()
                    if value:
                        pool.free_swim_price = value
                        updated = True

                # 주소
                if '주소' in row_dict and row_dict['주소']:
                    pool.address = str(row_dict['주소']).strip()
                    updated = True

                # 전화번호
                if '전화번호' in row_dict and row_dict['전화번호']:
                    pool.phone = str(row_dict['전화번호']).strip()
                    updated = True

                # 웹사이트
                if '웹사이트' in row_dict and row_dict['웹사이트']:
                    pool.url = str(row_dict['웹사이트']).strip()
                    updated = True

                if updated:
                    updated_count += 1

            except Exception as e:
                error_rows.append({
                    "row": row_num,
                    "error": str(e)
                })

        # 변경사항 커밋
        db.commit()

        return {
            "status": "success",
            "updated_count": updated_count,
            "total_rows": ws.max_row - 1,  # 헤더 제외
            "errors": error_rows if error_rows else None
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Excel 처리 중 오류 발생: {str(e)}")
